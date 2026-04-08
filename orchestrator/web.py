#!/usr/bin/env python3
"""
CLI Orchestrator — Web Viewer with plugin-based Harness architecture.

Usage:
    python3 orchestrator/web.py --port 30001
"""

import argparse
import json
import os
import uuid
from pathlib import Path

from dotenv import load_dotenv
from flask import Flask, render_template, request, jsonify

load_dotenv(Path(__file__).parent / ".env")

from harness import HarnessSession
from harness.engine import HarnessEngine
from plugins.excel import ExcelPlugin
from plugins.excel_com import ExcelComPlugin

app = Flask(__name__)
app.secret_key = os.urandom(24)

UPLOAD_DIR = Path(__file__).parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

# Plugin registry — add new plugins here
PLUGINS = {
    "excel": ExcelPlugin,
    "excel_com": ExcelComPlugin,
}

# Active sessions
sessions: dict[str, dict] = {}

# Persist recent files list
RECENT_FILE = Path(__file__).parent / "uploads" / ".recent.json"


def _detect_plugin(filename: str) -> str:
    """Detect plugin from file extension."""
    ext = Path(filename).suffix.lower()
    if ext in (".xlsx", ".xls", ".xlsm"):
        return "excel"
    return "excel"  # Default fallback


def load_recent_files() -> list[dict]:
    if RECENT_FILE.exists():
        try:
            return json.loads(RECENT_FILE.read_text())
        except Exception:
            pass
    return []


def save_recent_file(session_id: str, file_name: str, file_path: str, summary: dict):
    recents = load_recent_files()
    recents = [r for r in recents if r.get("file_path") != file_path]
    recents.insert(0, {
        "session_id": session_id,
        "file_name": file_name,
        "file_path": file_path,
        "summary": summary,
        "timestamp": __import__("datetime").datetime.now().isoformat(),
    })
    recents = recents[:20]
    RECENT_FILE.write_text(json.dumps(recents, ensure_ascii=False, indent=2))


def _create_session(file_path: str, file_name: str) -> tuple[str, dict, dict]:
    """Create a harness session and return (session_id, summary, session_dict)."""
    plugin_name = _detect_plugin(file_name)
    plugin = PLUGINS[plugin_name]()

    session_id = uuid.uuid4().hex[:12]
    summary = plugin.get_file_summary(file_path)

    harness = HarnessSession.create(
        plugin=plugin,
        context_path=file_path,
        provider="openai",
        model="gpt-5.4",
        project_root=Path(__file__).parent.parent,
    )

    sess = {
        "harness": harness,
        "engine": None,
        "file_path": file_path,
        "file_name": file_name,
        "summary": summary,
        "plugin_name": plugin_name,
    }
    sessions[session_id] = sess
    return session_id, summary, sess


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/recent", methods=["GET"])
def recent_files():
    recents = load_recent_files()
    valid = [r for r in recents if Path(r["file_path"]).exists()]
    return jsonify(valid)


@app.route("/reopen", methods=["POST"])
def reopen():
    data = request.json
    file_path = data.get("file_path", "")

    if not file_path or not Path(file_path).exists():
        return jsonify({"error": "File not found"}), 404

    file_name = Path(file_path).name
    if "_" in file_name and len(file_name.split("_")[0]) == 12:
        file_name = "_".join(file_name.split("_")[1:])

    session_id, summary, _ = _create_session(file_path, file_name)

    return jsonify({
        "session_id": session_id,
        "file_name": file_name,
        "summary": summary,
    })


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No filename"}), 400

    plugin_name = _detect_plugin(file.filename)
    plugin = PLUGINS[plugin_name]()

    tmp_id = uuid.uuid4().hex[:12]
    filename = f"{tmp_id}_{file.filename}"
    filepath = UPLOAD_DIR / filename
    file.save(str(filepath))

    ok, err = plugin.validate_file(str(filepath))
    if not ok:
        filepath.unlink(missing_ok=True)
        return jsonify({"error": err}), 400

    session_id, summary, _ = _create_session(str(filepath), file.filename)
    save_recent_file(session_id, file.filename, str(filepath), summary)

    return jsonify({
        "session_id": session_id,
        "file_name": file.filename,
        "summary": summary,
    })


@app.route("/chat", methods=["POST"])
def chat():
    data = request.json
    session_id = data.get("session_id")
    message = data.get("message", "").strip()
    provider = data.get("provider", "openai")
    model = data.get("model")

    if not session_id or session_id not in sessions:
        return jsonify({"error": "Invalid session"}), 400
    if not message:
        return jsonify({"error": "Empty message"}), 400

    sess = sessions[session_id]
    harness = sess["harness"]

    if not model:
        model = "gpt-5.4" if provider == "openai" else "gemini-3.1-pro-preview"

    if harness.provider != provider or harness.model != model:
        harness.provider = provider
        harness.model = model
        sess["engine"] = None

    if sess["engine"] is None:
        sess["engine"] = HarnessEngine(harness)

    try:
        response = sess["engine"].chat(message)
        return jsonify({"response": response})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/preview", methods=["POST"])
def preview():
    """Preview sheet data — delegates to plugin-specific logic."""
    data = request.json
    session_id = data.get("session_id")
    sheet_name = data.get("sheet", "")

    if not session_id or session_id not in sessions:
        return jsonify({"error": "Invalid session"}), 400
    if not sheet_name:
        return jsonify({"error": "Specify sheet"}), 400

    sess = sessions[session_id]

    # Excel-specific preview (kept for backward compat)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(sess["file_path"], read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return jsonify({"error": f"Sheet '{sheet_name}' not found"}), 400

        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(max_row=30, max_col=20, values_only=True):
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append(None)
                elif isinstance(cell, (int, float, bool)):
                    row_data.append(cell)
                else:
                    row_data.append(str(cell))
            rows.append(row_data)

        total_rows = ws.max_row or 0
        total_cols = ws.max_column or 0
        wb.close()

        return jsonify({
            "sheet": sheet_name,
            "rows": rows,
            "row_count": len(rows),
            "total_rows": total_rows,
            "total_cols": total_cols,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/session-info", methods=["POST"])
def session_info():
    data = request.json
    session_id = data.get("session_id")

    if not session_id or session_id not in sessions:
        return jsonify({"error": "Invalid session"}), 400

    sess = sessions[session_id]
    harness = sess["harness"]

    return jsonify({
        "session_id": session_id,
        "plugin": harness.plugin_name,
        "provider": harness.provider,
        "model": harness.model,
        "tools": harness.tools.tool_names(),
        "skills": harness.skills.get_skill_names(),
        "message_count": len(harness.messages),
    })


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=5000)
    parser.add_argument("--host", default="0.0.0.0")
    args = parser.parse_args()

    print(f"Harness Web UI: http://localhost:{args.port}")
    print(f"Plugins: {', '.join(PLUGINS.keys())}")
    app.run(host=args.host, port=args.port, debug=False)

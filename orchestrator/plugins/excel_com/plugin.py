"""Excel COM plugin — delegates heavy operations to Windows COM bridge server.

Architecture:
    Linux (harness)                    Windows (server.py)
    ├── Read: openpyxl (local, fast)
    ├── Write: → HTTP → ──────────────→ COM → Excel.exe
    ├── Pivot: → HTTP → ──────────────→ COM → Excel.exe
    ├── Chart: → HTTP → ──────────────→ COM → Excel.exe
    └── Calc:  → HTTP → ──────────────→ COM → Excel.exe

Setup:
    1. Windows: pip install pywin32 flask
    2. Windows: python server.py --port 8765
    3. Linux:   Set EXCEL_COM_SERVER=http://<windows-ip>:8765
    4. Linux:   Use --plugin excel_com in harness
"""

from __future__ import annotations

import json
import os
from pathlib import Path

import openpyxl
import requests

from harness.hooks import HookDecision, HookResult
from harness.tools import ToolDef, ToolRegistry
from plugins.base import HarnessPlugin


class ExcelComPlugin(HarnessPlugin):
    """Excel plugin with Windows COM bridge for full Excel engine support.

    Read operations use local openpyxl (fast).
    Write/pivot/chart/calc operations go to Windows COM bridge server.
    """

    def __init__(self):
        self.server_url = os.environ.get("EXCEL_COM_SERVER", "http://localhost:8765")
        self._com_session_id = None

    @property
    def name(self) -> str:
        return "excel_com"

    @property
    def description(self) -> str:
        return "Excel with Windows COM bridge (pivot, chart, formula calc)"

    def find_cli(self, project_root: Path) -> str:
        candidates = [
            project_root / "target" / "debug" / "excel-cli",
            project_root / "target" / "release" / "excel-cli",
        ]
        for c in candidates:
            if c.exists():
                return str(c)
        return ""

    def validate_file(self, file_path: str) -> tuple[bool, str]:
        p = Path(file_path)
        if not p.exists():
            return False, f"File not found: {file_path}"
        if p.suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
            return False, f"Unsupported file type: {p.suffix}"

        # Check COM server connectivity
        try:
            resp = requests.get(f"{self.server_url}/health", timeout=3)
            if resp.status_code != 200:
                return True, f"Warning: COM server not responding at {self.server_url}"
        except requests.ConnectionError:
            return True, f"Warning: COM server not reachable at {self.server_url}. Read-only mode."

        return True, ""

    def get_file_summary(self, file_path: str) -> dict:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets_info = [
                {
                    "name": name,
                    "rows": wb[name].max_row or 0,
                    "cols": wb[name].max_column or 0,
                    "visible": wb[name].sheet_state == "visible",
                }
                for name in wb.sheetnames
            ]
            summary = {
                "file": Path(file_path).name,
                "size_bytes": os.path.getsize(file_path),
                "sheet_count": len(wb.sheetnames),
                "sheets": sheets_info,
            }
            wb.close()
            return summary
        except Exception as e:
            return {"error": str(e), "sheet_count": 0, "sheets": []}

    def build_system_prompt(self, file_path: str, file_summary: dict,
                            skill_names: list[str], tool_names: list[str]) -> str:
        sheets_desc = ""
        if file_summary.get("sheets"):
            sheets_desc = ", ".join(
                f"{s['name']}({s.get('rows', 0)}x{s.get('cols', 0)})"
                for s in file_summary["sheets"]
            )

        return (
            "You are an Excel data analyst with FULL Excel engine access.\n"
            "You can read, write, create pivot tables, charts, recalculate formulas, and export PDFs.\n\n"
            f"## Current File\nPath: {file_path}\n"
            f"Sheets: {sheets_desc}\n"
            f"Size: {file_summary.get('size_bytes', 0):,} bytes\n\n"
            "## Guidelines\n"
            "- Use excel_summarize first to understand structure\n"
            "- Use excel_com_* tools for write/pivot/chart/calc (these use Windows Excel engine)\n"
            "- Use local read tools for fast data reading\n"
            "- For pivot tables: specify source range, row fields, value fields\n"
            "- For charts: specify source range and chart type\n"
            "- After writes/formulas, use excel_com_recalculate to refresh\n"
            "- Confirm destructive operations with the user\n"
            "- Respond in the same language the user uses"
        )

    def get_default_hooks(self) -> list[dict]:
        def write_guard(tool_name, tool_input, _output):
            if "write" in tool_name or "set" in tool_name:
                return HookResult(
                    decision=HookDecision.ALLOW,
                    extra_context="[SAFETY] This modifies the Excel file via Windows COM.",
                )
            return HookResult()

        return [
            {
                "event": "PreToolUse",
                "fn": write_guard,
                "matcher": "excel_com_*",
                "description": "Warn on COM write operations",
            }
        ]

    def register_tools(self, registry: ToolRegistry) -> None:
        """Register both local read tools and COM bridge tools."""
        self._register_local_read_tools(registry)
        self._register_com_tools(registry)

    # ═══════════════════════════════════════════
    # LOCAL READ TOOLS (openpyxl — fast)
    # ═══════════════════════════════════════════

    def _register_local_read_tools(self, registry: ToolRegistry):
        tools = [
            ToolDef("excel_summarize",
                "Get workbook summary: sheets, dimensions, size. Use this FIRST.",
                {"type": "object", "properties": {}, "required": []},
                self._read_summarize),
            ToolDef("excel_sheet_list",
                "List all sheets.",
                {"type": "object", "properties": {}, "required": []},
                self._read_sheet_list),
            ToolDef("excel_range_read",
                "Read cell values. Use 'Sheet1!A1:D10'. Max 50 rows.",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Range like 'Sheet1!A1:D10'"},
                }, "required": ["range"]},
                self._read_range),
            ToolDef("excel_cell_read",
                "Read a single cell.",
                {"type": "object", "properties": {
                    "cell": {"type": "string", "description": "Cell like 'Sheet1!B5'"},
                }, "required": ["cell"]},
                self._read_cell),
            ToolDef("excel_find",
                "Search for a value across sheets.",
                {"type": "object", "properties": {
                    "query": {"type": "string", "description": "Value to search"},
                    "sheet": {"type": "string", "description": "Optional: limit to sheet"},
                }, "required": ["query"]},
                self._read_find),
            ToolDef("excel_export_csv",
                "Export sheet as CSV.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "max_rows": {"type": "string", "description": "Max rows (default: 100)"},
                }, "required": ["sheet"]},
                self._read_export_csv),
        ]
        for t in tools:
            registry.register(t)

    # ═══════════════════════════════════════════
    # COM BRIDGE TOOLS (via Windows HTTP server)
    # ═══════════════════════════════════════════

    def _register_com_tools(self, registry: ToolRegistry):
        tools = [
            # Cell/Range write
            ToolDef("excel_com_open",
                "Open file in Windows Excel for editing. REQUIRED before write/pivot/chart.",
                {"type": "object", "properties": {
                    "visible": {"type": "string", "description": "'true' to show Excel window"},
                }, "required": []},
                self._com_open),
            ToolDef("excel_com_save",
                "Save the open workbook.",
                {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Optional: save-as path"},
                }, "required": []},
                self._com_save),
            ToolDef("excel_com_close",
                "Close workbook and Excel.",
                {"type": "object", "properties": {
                    "save": {"type": "string", "description": "'true' to save before closing"},
                }, "required": []},
                self._com_close),
            ToolDef("excel_com_set_cell",
                "Write a value to a cell. CONFIRM with user.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell": {"type": "string", "description": "Cell like 'C7'"},
                    "value": {"type": "string", "description": "Value to write"},
                }, "required": ["sheet", "cell", "value"]},
                self._com_set_cell),
            ToolDef("excel_com_set_range",
                "Write 2D array to a range. CONFIRM with user.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "range": {"type": "string", "description": "Range like 'A1:C3'"},
                    "data": {"type": "string", "description": "JSON 2D array"},
                }, "required": ["sheet", "range", "data"]},
                self._com_set_range),
            ToolDef("excel_com_get_cell",
                "Read cell with formula and formatting via Excel engine.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell": {"type": "string", "description": "Cell reference"},
                }, "required": ["sheet", "cell"]},
                self._com_get_cell),

            # Formula / Calc
            ToolDef("excel_com_recalculate",
                "Recalculate ALL formulas in workbook.",
                {"type": "object", "properties": {}, "required": []},
                self._com_recalculate),
            ToolDef("excel_com_set_formula",
                "Write a formula to a cell.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell": {"type": "string", "description": "Cell reference"},
                    "formula": {"type": "string", "description": "Formula like '=SUM(A1:A10)'"},
                }, "required": ["sheet", "cell", "formula"]},
                self._com_set_formula),
            ToolDef("excel_com_formula_result",
                "Get calculated result of a formula cell.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "cell": {"type": "string", "description": "Cell reference"},
                }, "required": ["sheet", "cell"]},
                self._com_formula_result),

            # Pivot
            ToolDef("excel_com_create_pivot",
                "Create a pivot table. Specify source data, row fields, and value fields.",
                {"type": "object", "properties": {
                    "source_sheet": {"type": "string", "description": "Sheet with source data"},
                    "source_range": {"type": "string", "description": "Source range like 'A1:E100'"},
                    "dest_sheet": {"type": "string", "description": "Destination sheet (created if missing)"},
                    "name": {"type": "string", "description": "Pivot table name"},
                    "row_fields": {"type": "string", "description": "JSON array of row field names"},
                    "value_fields": {"type": "string", "description": "JSON array of {name, function} objects"},
                    "col_fields": {"type": "string", "description": "Optional: JSON array of column fields"},
                }, "required": ["source_sheet", "source_range", "row_fields", "value_fields"]},
                self._com_create_pivot),
            ToolDef("excel_com_refresh_pivot",
                "Refresh a pivot table.",
                {"type": "object", "properties": {
                    "name": {"type": "string", "description": "Pivot table name"},
                }, "required": ["name"]},
                self._com_refresh_pivot),
            ToolDef("excel_com_list_pivots",
                "List all pivot tables.",
                {"type": "object", "properties": {}, "required": []},
                self._com_list_pivots),

            # Chart
            ToolDef("excel_com_create_chart",
                "Create a chart from data range.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet with data"},
                    "source_range": {"type": "string", "description": "Data range"},
                    "chart_type": {"type": "string", "description": "column, bar, line, pie, scatter, area"},
                    "title": {"type": "string", "description": "Chart title"},
                }, "required": ["sheet", "source_range"]},
                self._com_create_chart),

            # Export
            ToolDef("excel_com_export_pdf",
                "Export workbook to PDF.",
                {"type": "object", "properties": {
                    "output": {"type": "string", "description": "Output PDF path"},
                    "sheets": {"type": "string", "description": "Optional: JSON array of sheet names"},
                }, "required": ["output"]},
                self._com_export_pdf),

            # VBA
            ToolDef("excel_com_run_macro",
                "Run a VBA macro. CONFIRM with user.",
                {"type": "object", "properties": {
                    "name": {"type": "string", "description": "Macro name"},
                }, "required": ["name"]},
                self._com_run_macro),
        ]
        for t in tools:
            registry.register(t)

    # ── COM HTTP helpers ──

    def _com_request(self, endpoint: str, data: dict | None = None) -> str:
        """Send request to Windows COM bridge server."""
        data = data or {}
        if self._com_session_id:
            data.setdefault("session_id", self._com_session_id)

        try:
            resp = requests.post(
                f"{self.server_url}{endpoint}",
                json=data,
                timeout=120,
            )
            return resp.text
        except requests.ConnectionError:
            return json.dumps({
                "error": f"COM server not reachable at {self.server_url}. "
                         "Start server.py on Windows first."
            })
        except requests.Timeout:
            return json.dumps({"error": "COM server request timed out"})

    # ── Local read implementations ──

    def _open_readonly(self, fp: str):
        return openpyxl.load_workbook(fp, read_only=True, data_only=True)

    def _parse_ref(self, ref: str, default_sheet: str = "") -> tuple[str, str]:
        if "!" in ref:
            sheet, rng = ref.split("!", 1)
            return sheet, rng
        return default_sheet, ref

    def _read_summarize(self, args: dict, fp: str) -> str:
        return json.dumps(self.get_file_summary(fp), ensure_ascii=False)

    def _read_sheet_list(self, args: dict, fp: str) -> str:
        wb = self._open_readonly(fp)
        sheets = [{"name": n, "index": i, "visible": wb[n].sheet_state == "visible"}
                  for i, n in enumerate(wb.sheetnames)]
        wb.close()
        return json.dumps(sheets, ensure_ascii=False)

    def _read_range(self, args: dict, fp: str) -> str:
        range_str = args.get("range", "A1:A1")
        wb = self._open_readonly(fp)
        sheet, rng = self._parse_ref(range_str, wb.sheetnames[0])
        if sheet not in wb.sheetnames:
            wb.close()
            return json.dumps({"error": f"Sheet '{sheet}' not found"})
        ws = wb[sheet]
        start = ws[rng[0] if ':' not in rng else rng]
        # Handle calamine offset
        range_start = ws.min_row or 1
        rows = []
        for i, row in enumerate(ws[rng]):
            if i >= 50:
                break
            rows.append([
                cell.value if isinstance(cell.value, (int, float, bool)) or cell.value is None
                else str(cell.value) for cell in row
            ])
        wb.close()
        return json.dumps({"sheet": sheet, "range": range_str, "rows": rows, "row_count": len(rows)}, ensure_ascii=False)

    def _read_cell(self, args: dict, fp: str) -> str:
        ref = args.get("cell", "A1")
        wb = self._open_readonly(fp)
        sheet, cell = self._parse_ref(ref, wb.sheetnames[0])
        if sheet not in wb.sheetnames:
            wb.close()
            return json.dumps({"error": f"Sheet '{sheet}' not found"})
        val = wb[sheet][cell].value
        wb.close()
        return json.dumps({"cell": ref, "value": val}, ensure_ascii=False, default=str)

    def _read_find(self, args: dict, fp: str) -> str:
        query = str(args.get("query", "")).lower()
        target = args.get("sheet", "")
        wb = self._open_readonly(fp)
        matches = []
        for name in wb.sheetnames:
            if target and name != target:
                continue
            ws = wb[name]
            for row in ws.iter_rows(max_row=min(ws.max_row or 0, 1000), values_only=False):
                for cell in row:
                    if cell.value is not None and query in str(cell.value).lower():
                        matches.append({"sheet": name, "cell": cell.coordinate, "value": str(cell.value)[:200]})
                        if len(matches) >= 20:
                            break
                if len(matches) >= 20:
                    break
            if len(matches) >= 20:
                break
        wb.close()
        return json.dumps({"query": query, "matches": matches, "count": len(matches)}, ensure_ascii=False)

    def _read_export_csv(self, args: dict, fp: str) -> str:
        sheet = args.get("sheet", "")
        max_rows = int(args.get("max_rows", "100"))
        wb = self._open_readonly(fp)
        if sheet not in wb.sheetnames:
            wb.close()
            return json.dumps({"error": f"Sheet '{sheet}' not found"})
        ws = wb[sheet]
        lines = []
        for row in ws.iter_rows(max_row=max_rows, values_only=True):
            cells = []
            for v in row:
                if v is None:
                    cells.append("")
                elif isinstance(v, str) and ("," in v or '"' in v or "\n" in v):
                    cells.append(f'"{v.replace(chr(34), chr(34)+chr(34))}"')
                else:
                    cells.append(str(v))
            lines.append(",".join(cells))
        wb.close()
        return "\n".join(lines)

    # ── COM tool implementations ──

    def _com_open(self, args: dict, fp: str) -> str:
        visible = args.get("visible", "false") == "true"
        result = self._com_request("/open", {"path": fp, "visible": visible})
        try:
            data = json.loads(result)
            if "session_id" in data:
                self._com_session_id = data["session_id"]
        except json.JSONDecodeError:
            pass
        return result

    def _com_save(self, args: dict, fp: str) -> str:
        return self._com_request("/save", {"path": args.get("path")})

    def _com_close(self, args: dict, fp: str) -> str:
        save = args.get("save", "false") == "true"
        result = self._com_request("/close", {"save": save})
        self._com_session_id = None
        return result

    def _com_set_cell(self, args: dict, fp: str) -> str:
        value = args["value"]
        try:
            value = float(value) if "." in value else int(value)
        except (ValueError, TypeError):
            pass
        return self._com_request("/cell/set", {
            "sheet": args["sheet"], "cell": args["cell"], "value": value
        })

    def _com_set_range(self, args: dict, fp: str) -> str:
        data = json.loads(args["data"])
        return self._com_request("/range/set", {
            "sheet": args["sheet"], "range": args["range"], "data": data
        })

    def _com_get_cell(self, args: dict, fp: str) -> str:
        return self._com_request("/cell/get", {
            "sheet": args["sheet"], "cell": args["cell"]
        })

    def _com_recalculate(self, args: dict, fp: str) -> str:
        return self._com_request("/recalculate")

    def _com_set_formula(self, args: dict, fp: str) -> str:
        return self._com_request("/formula/set", {
            "sheet": args["sheet"], "cell": args["cell"], "formula": args["formula"]
        })

    def _com_formula_result(self, args: dict, fp: str) -> str:
        return self._com_request("/formula/result", {
            "sheet": args["sheet"], "cell": args["cell"]
        })

    def _com_create_pivot(self, args: dict, fp: str) -> str:
        return self._com_request("/pivot/create", {
            "source_sheet": args["source_sheet"],
            "source_range": args["source_range"],
            "dest_sheet": args.get("dest_sheet", "PivotResult"),
            "dest_cell": args.get("dest_cell", "A1"),
            "name": args.get("name", "PivotTable1"),
            "row_fields": json.loads(args["row_fields"]),
            "value_fields": json.loads(args["value_fields"]),
            "col_fields": json.loads(args["col_fields"]) if args.get("col_fields") else None,
        })

    def _com_refresh_pivot(self, args: dict, fp: str) -> str:
        return self._com_request("/pivot/refresh", {"name": args["name"]})

    def _com_list_pivots(self, args: dict, fp: str) -> str:
        return self._com_request("/pivot/list")

    def _com_create_chart(self, args: dict, fp: str) -> str:
        return self._com_request("/chart/create", {
            "sheet": args["sheet"],
            "source_range": args["source_range"],
            "chart_type": args.get("chart_type", "column"),
            "title": args.get("title", ""),
        })

    def _com_export_pdf(self, args: dict, fp: str) -> str:
        sheets = json.loads(args["sheets"]) if args.get("sheets") else None
        return self._com_request("/export/pdf", {
            "output": args["output"], "sheets": sheets
        })

    def _com_run_macro(self, args: dict, fp: str) -> str:
        return self._com_request("/macro/run", {"name": args["name"]})

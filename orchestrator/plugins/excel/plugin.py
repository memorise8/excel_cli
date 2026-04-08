"""Excel plugin — adapts excel-cli + openpyxl to the harness framework."""

from __future__ import annotations

import json
import os
from pathlib import Path

import openpyxl

from harness.hooks import HookDecision, HookResult
from harness.tools import ToolDef, ToolRegistry
from plugins.base import HarnessPlugin


class ExcelPlugin(HarnessPlugin):
    """Plugin for Excel file analysis and manipulation.

    - Read operations: openpyxl read_only (fast, large file safe)
    - Write operations: excel-cli subprocess (full Rust engine)
    """

    @property
    def name(self) -> str:
        return "excel"

    @property
    def description(self) -> str:
        return "Excel workbook analysis and manipulation via excel-cli"

    def find_cli(self, project_root: Path) -> str:
        candidates = [
            project_root / "target" / "debug" / "excel-cli",
            project_root / "target" / "release" / "excel-cli",
        ]
        for c in candidates:
            if c.exists():
                return str(c)
        return ""

    def validate_file(self, file_path: str) -> tuple[bool, str]:
        p = Path(file_path)
        if not p.exists():
            return False, f"File not found: {file_path}"
        if p.suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
            return False, f"Unsupported file type: {p.suffix}"
        return True, ""

    def get_file_summary(self, file_path: str) -> dict:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheets_info = [
                {
                    "name": name,
                    "rows": wb[name].max_row or 0,
                    "cols": wb[name].max_column or 0,
                    "visible": wb[name].sheet_state == "visible",
                }
                for name in wb.sheetnames
            ]
            summary = {
                "file": Path(file_path).name,
                "size_bytes": os.path.getsize(file_path),
                "sheet_count": len(wb.sheetnames),
                "sheets": sheets_info,
            }
            wb.close()
            return summary
        except Exception as e:
            return {"error": str(e), "sheet_count": 0, "sheets": []}

    def build_system_prompt(
        self,
        file_path: str,
        file_summary: dict,
        skill_names: list[str],
        tool_names: list[str],
    ) -> str:
        parts = [
            "You are an Excel data analyst assistant. "
            "You help users understand and analyze Excel files using available tools."
        ]

        if file_summary and file_summary.get("sheets"):
            sheets_desc = ", ".join(
                f"{s.get('name', '?')}({s.get('rows', 0)}x{s.get('cols', 0)})"
                for s in file_summary["sheets"]
            )
            parts.append(f"\n## Current File\nPath: {file_path}")
            parts.append(f"Sheets: {sheets_desc}")
            parts.append(f"Size: {file_summary.get('size_bytes', 0):,} bytes")

        parts.append(
            "\n## Excel Guidelines\n"
            "- First use excel_summarize to understand file structure.\n"
            "- Use excel_range_read for specific data ranges.\n"
            "- For large files, always limit range (max 50 rows) to avoid timeouts.\n"
            "- Provide clear, concise answers based on actual data.\n"
            "- For write operations, confirm with the user before making changes."
        )

        parts.append(
            "\n## Cloud Operations (Microsoft Graph API)\n"
            "- Use excel_cloud_auth_status to check if authenticated.\n"
            "- Workflow: upload file → get item_id → perform cloud ops → download result.\n"
            "- Use excel_cloud_range_read with with_format='true' to read formatting.\n"
            "- Use excel_cloud_calc to recalculate formulas after writing.\n"
            "- Cloud operations require 'excel-cli auth login' first."
        )

        return "\n".join(parts)

    def get_default_hooks(self) -> list[dict]:
        def write_guard(tool_name: str, tool_input: dict | None, _output: str):
            if tool_name in ("excel_range_write",):
                return HookResult(
                    decision=HookDecision.ALLOW,
                    extra_context="[SAFETY] This is a write operation. Data will be modified.",
                )
            return HookResult()

        return [
            {
                "event": "PreToolUse",
                "fn": write_guard,
                "matcher": "excel_range_write",
                "description": "Warn on Excel write operations",
            },
        ]

    def register_tools(self, registry: ToolRegistry) -> None:
        self._register_read_tools(registry)
        self._register_cli_tools(registry)
        self._register_cloud_tools(registry)

    # ═══════════════════════════════════════════
    # READ TOOLS (openpyxl — fast, large file safe)
    # ═══════════════════════════════════════════

    def _register_read_tools(self, registry: ToolRegistry) -> None:
        read_tools = [
            ToolDef("excel_summarize",
                "Get workbook summary: sheets, dimensions, file size. Use this FIRST.",
                {"type": "object", "properties": {}, "required": []},
                self._read_summarize),
            ToolDef("excel_sheet_list",
                "List all sheets with names and visibility.",
                {"type": "object", "properties": {}, "required": []},
                self._read_sheet_list),
            ToolDef("excel_range_read",
                "Read cell values from a range. Use 'Sheet1!A1:D10'. Max 50 rows for large files.",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Range like 'Sheet1!A1:D10'"},
                }, "required": ["range"]},
                self._read_range),
            ToolDef("excel_cell_read",
                "Read a single cell value.",
                {"type": "object", "properties": {
                    "cell": {"type": "string", "description": "Cell like 'Sheet1!B5'"},
                }, "required": ["cell"]},
                self._read_cell),
            ToolDef("excel_find",
                "Search for a value across sheets. Returns matching cells.",
                {"type": "object", "properties": {
                    "query": {"type": "string", "description": "Value to search for"},
                    "sheet": {"type": "string", "description": "Optional: limit to this sheet"},
                }, "required": ["query"]},
                self._read_find),
            ToolDef("excel_formula_read",
                "Read formula from a cell (formula text + cached value).",
                {"type": "object", "properties": {
                    "cell": {"type": "string", "description": "Cell like 'Sheet1!B10'"},
                }, "required": ["cell"]},
                self._read_formula),
            ToolDef("excel_export_csv",
                "Export a sheet as CSV text.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "max_rows": {"type": "string", "description": "Max rows (default: 100)"},
                }, "required": ["sheet"]},
                self._read_export_csv),
        ]
        for t in read_tools:
            registry.register(t)

    # ── openpyxl helpers ──

    @staticmethod
    def _open_readonly(fp: str) -> openpyxl.Workbook:
        return openpyxl.load_workbook(fp, read_only=True, data_only=True)

    @staticmethod
    def _parse_ref(ref: str, default_sheet: str = "") -> tuple[str, str]:
        if "!" in ref:
            sheet, rng = ref.split("!", 1)
            return sheet, rng
        return default_sheet, ref

    def _read_summarize(self, args: dict, fp: str) -> str:
        wb = self._open_readonly(fp)
        sheets = [{"name": n, "rows": wb[n].max_row or 0, "cols": wb[n].max_column or 0,
                    "visible": wb[n].sheet_state == "visible"} for n in wb.sheetnames]
        result = {"file": Path(fp).name, "size_bytes": os.path.getsize(fp),
                  "sheet_count": len(wb.sheetnames), "sheets": sheets}
        wb.close()
        return json.dumps(result, ensure_ascii=False)

    def _read_sheet_list(self, args: dict, fp: str) -> str:
        wb = self._open_readonly(fp)
        sheets = [{"name": n, "index": i, "visible": wb[n].sheet_state == "visible"}
                  for i, n in enumerate(wb.sheetnames)]
        wb.close()
        return json.dumps(sheets, ensure_ascii=False)

    def _read_range(self, args: dict, fp: str) -> str:
        range_str = args.get("range", "A1:A1")
        wb = self._open_readonly(fp)
        sheet, rng = self._parse_ref(range_str, wb.sheetnames[0])
        if sheet not in wb.sheetnames:
            wb.close()
            return json.dumps({"error": f"Sheet '{sheet}' not found"})
        ws = wb[sheet]
        rows = []
        for i, row in enumerate(ws[rng]):
            if i >= 50:
                break
            rows.append([cell.value if isinstance(cell.value, (int, float, bool)) or cell.value is None
                         else str(cell.value) for cell in row])
        wb.close()
        return json.dumps({"sheet": sheet, "range": range_str, "rows": rows, "row_count": len(rows)}, ensure_ascii=False)

    def _read_cell(self, args: dict, fp: str) -> str:
        ref = args.get("cell", "A1")
        wb = self._open_readonly(fp)
        sheet, cell = self._parse_ref(ref, wb.sheetnames[0])
        if sheet not in wb.sheetnames:
            wb.close()
            return json.dumps({"error": f"Sheet '{sheet}' not found"})
        val = wb[sheet][cell].value
        wb.close()
        return json.dumps({"cell": ref, "value": val}, ensure_ascii=False, default=str)

    def _read_find(self, args: dict, fp: str) -> str:
        query = str(args.get("query", "")).lower()
        target = args.get("sheet", "")
        wb = self._open_readonly(fp)
        matches = []
        for name in wb.sheetnames:
            if target and name != target:
                continue
            ws = wb[name]
            for row in ws.iter_rows(max_row=min(ws.max_row or 0, 1000), values_only=False):
                for cell in row:
                    if cell.value is not None and query in str(cell.value).lower():
                        matches.append({"sheet": name, "cell": cell.coordinate, "value": str(cell.value)[:200]})
                        if len(matches) >= 20:
                            break
                if len(matches) >= 20:
                    break
            if len(matches) >= 20:
                break
        wb.close()
        return json.dumps({"query": query, "matches": matches, "count": len(matches)}, ensure_ascii=False)

    def _read_formula(self, args: dict, fp: str) -> str:
        ref = args.get("cell", "A1")
        wb1 = openpyxl.load_workbook(fp, read_only=True, data_only=False)
        sheet, cell = self._parse_ref(ref, wb1.sheetnames[0])
        if sheet not in wb1.sheetnames:
            wb1.close()
            return json.dumps({"error": f"Sheet '{sheet}' not found"})
        raw = wb1[sheet][cell].value
        wb1.close()
        is_formula = isinstance(raw, str) and raw.startswith("=")
        wb2 = self._open_readonly(fp)
        cached = wb2[sheet][cell].value
        wb2.close()
        return json.dumps({"cell": ref, "formula": raw if is_formula else None,
                           "cached_value": cached, "has_formula": is_formula}, ensure_ascii=False, default=str)

    def _read_export_csv(self, args: dict, fp: str) -> str:
        sheet = args.get("sheet", "")
        max_rows = int(args.get("max_rows", "100"))
        wb = self._open_readonly(fp)
        if sheet not in wb.sheetnames:
            wb.close()
            return json.dumps({"error": f"Sheet '{sheet}' not found"})
        ws = wb[sheet]
        lines = []
        for row in ws.iter_rows(max_row=max_rows, values_only=True):
            cells = []
            for v in row:
                if v is None:
                    cells.append("")
                elif isinstance(v, str) and ("," in v or '"' in v or "\n" in v):
                    cells.append(f'"{v.replace(chr(34), chr(34)+chr(34))}"')
                else:
                    cells.append(str(v))
            lines.append(",".join(cells))
        wb.close()
        return "\n".join(lines)

    # ═══════════════════════════════════════════
    # CLI TOOLS (excel-cli subprocess — full engine)
    # ═══════════════════════════════════════════

    def _register_cli_tools(self, registry: ToolRegistry) -> None:
        cli_tools = [
            # ── Range write ──
            ToolDef("excel_range_write",
                "Write values to a range. Data is JSON 2D array. CONFIRM with user first.",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Target range, e.g. 'Sheet1!A1:C3'"},
                    "data": {"type": "string", "description": "JSON 2D array, e.g. '[[1,2],[3,4]]'"},
                }, "required": ["range", "data"]},
                lambda args, fp: registry.run_cli("range", "write", fp, args["range"], "-d", args["data"])),
            ToolDef("excel_range_clear",
                "Clear values from a range. CONFIRM with user first.",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Range to clear"},
                }, "required": ["range"]},
                lambda args, fp: registry.run_cli("range", "clear", fp, args["range"])),

            # ── Sheet management ──
            ToolDef("excel_sheet_add",
                "Add a new sheet. CONFIRM with user first.",
                {"type": "object", "properties": {
                    "name": {"type": "string", "description": "New sheet name"},
                }, "required": ["name"]},
                lambda args, fp: registry.run_cli("sheet", "add", fp, args["name"])),
            ToolDef("excel_sheet_rename",
                "Rename a sheet. CONFIRM with user first.",
                {"type": "object", "properties": {
                    "old_name": {"type": "string", "description": "Current sheet name"},
                    "new_name": {"type": "string", "description": "New sheet name"},
                }, "required": ["old_name", "new_name"]},
                lambda args, fp: registry.run_cli("sheet", "rename", fp, args["old_name"], args["new_name"])),
            ToolDef("excel_sheet_delete",
                "Delete a sheet. CONFIRM with user first.",
                {"type": "object", "properties": {
                    "name": {"type": "string", "description": "Sheet name to delete"},
                }, "required": ["name"]},
                lambda args, fp: registry.run_cli("sheet", "delete", fp, args["name"])),

            # ── Formula ──
            ToolDef("excel_formula_write",
                "Write a formula to a cell. CONFIRM with user first.",
                {"type": "object", "properties": {
                    "cell": {"type": "string", "description": "Target cell, e.g. 'Sheet1!B10'"},
                    "formula": {"type": "string", "description": "Excel formula, e.g. '=SUM(B1:B9)'"},
                }, "required": ["cell", "formula"]},
                lambda args, fp: registry.run_cli("formula", "write", fp, args["cell"], "--formula", args["formula"])),
            ToolDef("excel_formula_list",
                "List all formulas in a sheet.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet name"},
                }, "required": ["sheet"]},
                lambda args, fp: registry.run_cli("formula", "list", fp, "--sheet", args["sheet"])),

            # ── Format ──
            ToolDef("excel_format_font",
                "Apply font formatting (bold, size, color, italic).",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Target range"},
                    "bold": {"type": "string", "description": "'true' to enable bold"},
                    "size": {"type": "string", "description": "Font size, e.g. '14'"},
                    "color": {"type": "string", "description": "Hex color, e.g. 'FF0000'"},
                    "name": {"type": "string", "description": "Font name, e.g. 'Arial'"},
                }, "required": ["range"]},
                self._cli_format_font),
            ToolDef("excel_format_fill",
                "Set cell background color.",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Target range"},
                    "color": {"type": "string", "description": "Hex color, e.g. '4472C4'"},
                }, "required": ["range", "color"]},
                lambda args, fp: registry.run_cli("format", "fill", fp, args["range"], "--color", args["color"])),
            ToolDef("excel_format_border",
                "Set cell borders.",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Target range"},
                    "style": {"type": "string", "description": "Border style: thin, medium, thick"},
                    "color": {"type": "string", "description": "Hex color"},
                }, "required": ["range"]},
                self._cli_format_border),
            ToolDef("excel_format_number",
                "Set number format (currency, percent, date, etc).",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Target range"},
                    "preset": {"type": "string", "description": "Preset: currency, percent, date, number"},
                    "format": {"type": "string", "description": "Custom format code, e.g. '#,##0.00'"},
                }, "required": ["range"]},
                self._cli_format_number),
            ToolDef("excel_format_align",
                "Set cell alignment.",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Target range"},
                    "horizontal": {"type": "string", "description": "left, center, right"},
                    "vertical": {"type": "string", "description": "top, center, bottom"},
                    "wrap": {"type": "string", "description": "'true' to enable text wrap"},
                }, "required": ["range"]},
                self._cli_format_align),

            # ── Table ──
            ToolDef("excel_table_create",
                "Create an Excel table from a range. CONFIRM with user first.",
                {"type": "object", "properties": {
                    "range": {"type": "string", "description": "Data range, e.g. 'Sheet1!A1:D10'"},
                    "name": {"type": "string", "description": "Table name"},
                    "has_headers": {"type": "string", "description": "'true' if first row is header"},
                }, "required": ["range", "name"]},
                self._cli_table_create),
            ToolDef("excel_table_list",
                "List all tables in the workbook.",
                {"type": "object", "properties": {}, "required": []},
                lambda args, fp: registry.run_cli("table", "list", fp)),
            ToolDef("excel_table_read",
                "Read data from a named table.",
                {"type": "object", "properties": {
                    "name": {"type": "string", "description": "Table name"},
                }, "required": ["name"]},
                lambda args, fp: registry.run_cli("table", "read", fp, "--name", args["name"])),

            # ── Export (via CLI) ──
            ToolDef("excel_export_json",
                "Export sheet as JSON file.",
                {"type": "object", "properties": {
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "output": {"type": "string", "description": "Output file path"},
                }, "required": ["sheet"]},
                self._cli_export_json),

            # ── Helpers ──
            ToolDef("excel_diff",
                "Compare two Excel files.",
                {"type": "object", "properties": {
                    "file2": {"type": "string", "description": "Second file to compare with"},
                }, "required": ["file2"]},
                lambda args, fp: registry.run_cli("+diff", fp, args["file2"])),
            ToolDef("excel_template",
                "Create workbook from template (blank, budget, tracker, sales).",
                {"type": "object", "properties": {
                    "template": {"type": "string", "description": "Template name or path"},
                    "output": {"type": "string", "description": "Output file path"},
                }, "required": ["template", "output"]},
                lambda args, fp: registry.run_cli("+template", args["template"], args["output"])),
        ]
        for t in cli_tools:
            registry.register(t)

        # Store registry ref for complex CLI tools
        self._registry = registry

    # ── Complex CLI tool implementations (need arg building) ──

    def _cli_format_font(self, args: dict, fp: str) -> str:
        cmd = ["format", "font", fp, args["range"]]
        if args.get("bold") == "true":
            cmd.append("--bold")
        if args.get("size"):
            cmd.extend(["--size", args["size"]])
        if args.get("color"):
            cmd.extend(["--color", args["color"]])
        if args.get("name"):
            cmd.extend(["--name", args["name"]])
        return self._registry.run_cli(*cmd)

    def _cli_format_border(self, args: dict, fp: str) -> str:
        cmd = ["format", "border", fp, args["range"]]
        if args.get("style"):
            cmd.extend(["--style", args["style"]])
        if args.get("color"):
            cmd.extend(["--color", args["color"]])
        return self._registry.run_cli(*cmd)

    def _cli_format_number(self, args: dict, fp: str) -> str:
        cmd = ["format", "number", fp, args["range"]]
        if args.get("preset"):
            cmd.extend(["--preset", args["preset"]])
        if args.get("format"):
            cmd.extend(["--format", args["format"]])
        return self._registry.run_cli(*cmd)

    def _cli_format_align(self, args: dict, fp: str) -> str:
        cmd = ["format", "align", fp, args["range"]]
        if args.get("horizontal"):
            cmd.extend(["--horizontal", args["horizontal"]])
        if args.get("vertical"):
            cmd.extend(["--vertical", args["vertical"]])
        if args.get("wrap") == "true":
            cmd.append("--wrap")
        return self._registry.run_cli(*cmd)

    def _cli_table_create(self, args: dict, fp: str) -> str:
        cmd = ["table", "create", fp, args["range"], "--name", args["name"]]
        if args.get("has_headers") == "true":
            cmd.append("--has-headers")
        return self._registry.run_cli(*cmd)

    def _cli_export_json(self, args: dict, fp: str) -> str:
        cmd = ["export", "json", fp, "--sheet", args["sheet"]]
        if args.get("output"):
            cmd.extend(["-o", args["output"]])
        return self._registry.run_cli(*cmd)

    # ═══════════════════════════════════════════
    # CLOUD TOOLS (Graph API via excel-cli --cloud)
    # ═══════════════════════════════════════════

    def _register_cloud_tools(self, registry: ToolRegistry) -> None:
        cloud_tools = [
            ToolDef("excel_cloud_auth_status",
                "Check Microsoft Graph API authentication status.",
                {"type": "object", "properties": {}, "required": []},
                lambda args, fp: registry.run_cli("auth", "status")),

            ToolDef("excel_cloud_upload",
                "Upload an Excel file to OneDrive. Returns item_id needed for all cloud operations.",
                {"type": "object", "properties": {
                    "file": {"type": "string", "description": "Local file path to upload (defaults to current file)"},
                }, "required": []},
                self._cloud_upload),

            ToolDef("excel_cloud_download",
                "Download an Excel file from OneDrive by item_id.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                    "output": {"type": "string", "description": "Local output file path"},
                }, "required": ["item_id", "output"]},
                lambda args, fp: registry.run_cli("file", "download", args["output"], "--item-id", args["item_id"])),

            ToolDef("excel_cloud_range_read",
                "Read range data from cloud Excel with optional formatting. Returns values, formulas, numberFormat, and format details.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                    "range": {"type": "string", "description": "Range like 'Sheet1!A1:D10'"},
                    "with_format": {"type": "string", "description": "'true' to include font/fill/border formatting"},
                }, "required": ["item_id", "range"]},
                self._cloud_range_read),

            ToolDef("excel_cloud_range_write",
                "Write values to a cloud Excel range. CONFIRM with user first.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                    "range": {"type": "string", "description": "Target range like 'Sheet1!A1:C3'"},
                    "data": {"type": "string", "description": "JSON 2D array of values"},
                }, "required": ["item_id", "range", "data"]},
                lambda args, fp: registry.run_cli("range", "write", fp, args["range"],
                    "-d", args["data"], "--cloud", "--item-id", args["item_id"])),

            ToolDef("excel_cloud_format_read",
                "Read complete formatting (font, fill, borders) from a cloud Excel range.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                    "range": {"type": "string", "description": "Range like 'Sheet1!A1:D10'"},
                }, "required": ["item_id", "range"]},
                self._cloud_format_read),

            ToolDef("excel_cloud_calc",
                "Trigger full workbook recalculation via Graph API.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                }, "required": ["item_id"]},
                lambda args, fp: registry.run_cli("calc", "now", "--cloud", "--item-id", args["item_id"])),

            ToolDef("excel_cloud_export_pdf",
                "Export workbook as PDF via Graph API.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                    "output": {"type": "string", "description": "Output PDF file path"},
                }, "required": ["item_id", "output"]},
                lambda args, fp: registry.run_cli("export", "pdf", fp, "--cloud",
                    "--item-id", args["item_id"], "-o", args["output"])),

            ToolDef("excel_cloud_chart_list",
                "List all charts in a worksheet via Graph API.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                }, "required": ["item_id", "sheet"]},
                lambda args, fp: registry.run_cli("chart", "list", "--cloud",
                    "--item-id", args["item_id"], "--sheet", args["sheet"])),

            ToolDef("excel_cloud_chart_create",
                "Create a chart via Graph API. CONFIRM with user first.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "type": {"type": "string", "description": "Chart type (e.g. ColumnClustered, Line, Pie)"},
                    "source": {"type": "string", "description": "Data range for the chart"},
                }, "required": ["item_id", "sheet", "type", "source"]},
                lambda args, fp: registry.run_cli("chart", "create", "--cloud",
                    "--item-id", args["item_id"], "--sheet", args["sheet"],
                    "--type", args["type"], "--source", args["source"])),

            ToolDef("excel_cloud_pivot_list",
                "List all pivot tables in a worksheet via Graph API.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                }, "required": ["item_id", "sheet"]},
                lambda args, fp: registry.run_cli("pivot", "list", "--cloud",
                    "--item-id", args["item_id"], "--sheet", args["sheet"])),

            ToolDef("excel_cloud_pivot_refresh",
                "Refresh a pivot table via Graph API.",
                {"type": "object", "properties": {
                    "item_id": {"type": "string", "description": "OneDrive item ID"},
                    "sheet": {"type": "string", "description": "Sheet name"},
                    "name": {"type": "string", "description": "Pivot table name"},
                }, "required": ["item_id", "sheet", "name"]},
                lambda args, fp: registry.run_cli("pivot", "refresh", "--cloud",
                    "--item-id", args["item_id"], "--sheet", args["sheet"],
                    "--name", args["name"])),
        ]
        for t in cloud_tools:
            registry.register(t)

        self._registry = registry

    def _cloud_upload(self, args: dict, fp: str) -> str:
        file_to_upload = args.get("file", fp)
        return self._registry.run_cli("file", "upload", file_to_upload)

    def _cloud_range_read(self, args: dict, fp: str) -> str:
        cmd = ["range", "read", fp, args["range"], "--cloud", "--item-id", args["item_id"]]
        if args.get("with_format") == "true":
            cmd.append("--with-format")
        return self._registry.run_cli(*cmd)

    def _cloud_format_read(self, args: dict, fp: str) -> str:
        return self._registry.run_cli(
            "range", "read", fp, args["range"],
            "--cloud", "--item-id", args["item_id"], "--with-format"
        )
